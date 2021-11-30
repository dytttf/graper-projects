from graper.spiders import *
from graper.utils import log
import js2py
import json
import datetime

import openpyxl
import tqdm

logger = log.get_logger(__file__)


class AppListSpider(Spider):
    """
        采集指定行业下的APP列表
    """

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        encrypt_js = open("encrypt.js", "r", encoding="utf8").read()
        self.js_context = js2py.EvalJs()
        self.js_context.execute(encrypt_js)

        self.data = {}

        self.downloader.proxy_enable = False

    def before_stop(self, **kwargs):
        with open("app_list.json", "w", encoding="utf8") as f:
            json.dump(self.data, f)

        return

    def encrypt(self, text):
        return self.js_context.encode(
            "ZGFwcH" + "JhZGFy" + self.js_context.encode(text)
        )

    def make_request(self, cat, page):
        params = f"page={page}&sgroup=max&featured=1&range=day&category={cat}&sort=user&order=desc&limit=26"
        url = "https://dappradar.com/v2/api/dapps?params={}".format(
            self.encrypt(params)
        )
        return Request(url, meta={"cat": cat, "page": page})

    def start_requests(self):
        category_list = [
            "games",
            "collectibles",
            "gambling",
        ]
        for cat in category_list:
            yield self.make_request(cat, 1)

    def parse(self, response: Response):
        meta = response.request.meta
        cat = meta["cat"]
        j_response = response.response.json()

        for item in j_response["dapps"]:
            self.data[item["id"]] = item

        page = j_response["page"]
        print(cat, page, len(self.data))

        if page < j_response["pageCount"]:
            yield self.make_request(cat, page + 1)

        return


class AppDetailSpider(Spider):
    """
        采集app详情
    """

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.data = {}
        self.downloader.proxy_enable = False

    def before_stop(self, **kwargs):
        with open("app_detail.json", "w", encoding="utf8") as f:
            json.dump(self.data, f)
        return

    def start_requests(self):
        with open("app_list.json", "r", encoding="utf8") as f:
            for k, v in json.load(f).items():
                yield Request(
                    f"https://dappradar.com/v2/api/dapp/{v['protocolSlug']}/{v['category']}/{v['slug']}",
                    meta=v,
                )

    def parse(self, response: Response):
        try:
            meta = response.request.meta
            j_response = response.response.json()

            self.data[meta["id"]] = {
                "detail": j_response,
            }
            yield Request(
                f"https://dappradar.com/v2/api/dapp/{meta['protocolSlug']}/{meta['category']}/{meta['slug']}/chart/all",
                meta=meta,
                callback=self.parse_chart,
            )
        except Exception as e:
            logger.exception(e)
            yield response.request
        return

    def parse_chart(self, response: Response):
        try:
            meta = response.request.meta
            j_response = response.response.json()

            self.data[meta["id"]].update(
                {"chart": j_response,}
            )
            assert len(self.data[meta["id"]]) == 2

            print(len(self.data))
        except Exception as e:
            logger.exception(e)
            yield response.request
        return

    def export_data(self):
        with open("app_detail.json", "r", encoding="utf8") as f:
            data = json.load(f)

        headers = [
            "date",
            "Users",
            "Volume",
            "Transactions",
            "category",
            "name",
            "app_id",
        ]

        values = []
        for item in tqdm.tqdm(data.values(), desc="reading", total=len(data)):
            chart = item["chart"]
            detail = item["detail"]
            user_list = volumn_list = trans_list = None
            for x in chart["series"]:
                if x["name"] == "Users":
                    user_list = x["data"]
                elif x["name"] == "Volume":
                    volumn_list = x["data"]
                elif x["name"] == "Transactions":
                    trans_list = x["data"]
            _values = []
            for date, user, volume, trans in zip(
                chart["xaxis"], user_list, volumn_list, trans_list
            ):
                date = datetime.datetime.fromtimestamp(date / 1000).strftime("%Y-%m-%d")
                _values.append(
                    [
                        date,
                        user,
                        volume,
                        trans,
                        detail["category"],
                        detail["name"],
                        detail["id"],
                    ]
                )

            _values.sort(key=lambda x: x[0])
            _values.reverse()

            if not _values:
                _values.append(
                    ["", 0, 0, 0, detail["category"], detail["name"], detail["id"]]
                )

            values.extend(_values)

        print(len(values))

        wb = openpyxl.Workbook(write_only=True)
        sheet = wb.create_sheet()
        sheet.append(headers)
        for v in values:
            sheet.append(v)

        wb.save("dappradar.xlsx")
        return


class IndustryOverViewSpider(Spider):
    """
        采集行业大盘图表
    """

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.data = {}
        self.downloader.proxy_enable = False

    def before_stop(self, **kwargs):
        with open("industry_overview.json", "w", encoding="utf8") as f:
            json.dump(self.data, f)
        return

    def start_requests(self):
        for url in [
            "https://dappradar.com/api/charts/users-activity/category/history/year?currency=USD",
            "https://dappradar.com/api/charts/users-activity/protocol/history/year?currency=USD",
            "https://dappradar.com/api/charts/transactions/category/history/year?currency=USD",
            "https://dappradar.com/api/charts/transactions/protocol/history/year?currency=USD",
            "https://dappradar.com/api/charts/volume/category/history/year?currency=USD",
            "https://dappradar.com/api/charts/volume/protocol/history/year?currency=USD",
        ]:
            yield Request(url)

    def parse(self, response: Response):
        try:
            url = response.request.url
            j_response = response.response.json()

            self.data[url] = j_response
            print(len(self.data))
        except Exception as e:
            logger.exception(e)
            yield response.request
        return

    def export_data(self):
        with open("industry_overview.json", "r", encoding="utf8") as f:
            data = json.load(f)

        headers = [
            "date",
            "name",
            "value",
        ]

        wb = openpyxl.Workbook(write_only=True)

        for url, j_response in data.items():
            title = url.replace("https://dappradar.com/api/charts/", "")
            title = "-".join(title.split("/")[:2])
            sheet = wb.create_sheet(title=title)
            sheet.append(headers)

            date_list = j_response["xaxis"]
            for item in j_response["series"]:
                name = item["name"]
                for date, value in zip(date_list, item["data"]):
                    date = datetime.datetime.fromtimestamp(date / 1000).strftime(
                        "%Y-%m-%d"
                    )
                    sheet.append([date, name, value])
        wb.save("dappradar_industry_overview.xlsx")
        return


if __name__ == "__main__":
    with IndustryOverViewSpider(pool_size=5) as my_spider:
        my_spider.run()
